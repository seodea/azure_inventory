import io
from typing import Optional

from fastapi import APIRouter, HTTPException, Response
from pydantic import BaseModel
from azure.mgmt.network import NetworkManagementClient
from azure.mgmt.compute import ComputeManagementClient
from azure.mgmt.appcontainers import ContainerAppsAPIClient
from azure.mgmt.containerregistry import ContainerRegistryManagementClient
from azure.mgmt.cosmosdb import CosmosDBManagementClient
from azure.mgmt.rdbms.postgresql_flexibleservers import PostgreSQLManagementClient
from azure.mgmt.search import SearchManagementClient
from azure.mgmt.cognitiveservices import CognitiveServicesManagementClient
from azure.mgmt.redis import RedisManagementClient
from azure.mgmt.containerservice import ContainerServiceClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from common import get_credential, normalize_region, ServicePrincipalConfig

router = APIRouter()

class ExportRequest(BaseModel):
    tenant_id: str
    client_id: str
    client_secret: str
    subscription_id: str
    region: Optional[str] = None
    services: list[str]

# 색상
C_BLUE   = "2563EB"
C_INDIGO = "4338CA"
C_TEAL   = "0D9488"
C_ORANGE = "EA580C"
C_GRAY   = "475569"

# ─── 헬퍼 ─────────────────────────────────────────────────────────────
def _header(ws, headers, color=C_BLUE):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

def _autofit(ws, min_w=10, max_w=52):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

def _link(ws, row, col, sheet, display):
    cell = ws.cell(row=row, column=col, value=display)
    cell.hyperlink = f"#{sheet}!A1"
    cell.font = Font(color=C_BLUE, underline="single", size=10)

def _match(item, region):
    region_norm = normalize_region(region)
    if not region_norm:
        return True
    return normalize_region(getattr(item, "location", None)) == region_norm

def _rg(rid):
    p = (rid or "").split("/")
    return p[4] if len(p) > 4 else ""

# ─── VNet/Subnet ──────────────────────────────────────────────────────
def _sheet_vnet_subnet(wb, credential, sub_id, region):
    from network import _iter_vnet_subnets, ServicePrincipalConfig as NC
    ws = wb.create_sheet("VNet_Subnet")
    _header(ws, ["SubscriptionId","ResourceGroup","VNetName","VNetLocation",
                 "VNetAddressSpace","SubnetName","SubnetAddressPrefix",
                 "SubnetNSG","SubnetRouteTable","PrivateEndpointPolicies","PrivateLinkServicePolicies"])
    cfg = NC(tenant_id="",client_id="",client_secret="",subscription_id=sub_id,region=region)
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    for item in _iter_vnet_subnets(client, cfg):
        ws.append(list(item.values()))
    _autofit(ws)

# ─── NSG (규칙 행으로 전개) ────────────────────────────────────────────
def _sheet_nsg(wb, credential, sub_id, region):
    ws = wb.create_sheet("NSG")
    _header(ws, ["Name","ResourceGroup","Location","RuleName","Priority","Direction",
                 "Access","Protocol","SrcPort","DstPort","SrcAddr","DstAddr","SubscriptionId"], color=C_TEAL)
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    for nsg in client.network_security_groups.list_all():
        if not _match(nsg, region): continue
        rg = _rg(nsg.id)
        rules = list(nsg.security_rules or []) + list(nsg.default_security_rules or [])
        if not rules:
            ws.append([nsg.name, rg, nsg.location, "(없음)","","","","","","","","",sub_id])
        else:
            for r in sorted(rules, key=lambda x: getattr(x,"priority",9999)):
                ws.append([nsg.name, rg, nsg.location, r.name or "",
                           getattr(r,"priority",""), str(getattr(r,"direction","") or ""),
                           str(getattr(r,"access","") or ""), str(getattr(r,"protocol","") or ""),
                           getattr(r,"source_port_range","") or "",
                           getattr(r,"destination_port_range","") or "",
                           getattr(r,"source_address_prefix","") or "",
                           getattr(r,"destination_address_prefix","") or "", sub_id])
    _autofit(ws)

# ─── Route Table ──────────────────────────────────────────────────────
def _sheet_route_table(wb, credential, sub_id, region):
    ws = wb.create_sheet("RouteTable")
    _header(ws, ["Name","ResourceGroup","Location","RouteName","AddressPrefix",
                 "NextHopType","NextHopIP","SubscriptionId"], color=C_TEAL)
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    for rt in client.route_tables.list_all():
        if not _match(rt, region): continue
        rg = _rg(rt.id)
        routes = list(rt.routes or [])
        if not routes:
            ws.append([rt.name, rg, rt.location, "(없음)","","","",sub_id])
        else:
            for r in routes:
                ws.append([rt.name, rg, rt.location, r.name or "",
                           getattr(r,"address_prefix","") or "",
                           str(getattr(r,"next_hop_type","") or ""),
                           getattr(r,"next_hop_ip_address","") or "", sub_id])
    _autofit(ws)

# ─── Public IP ────────────────────────────────────────────────────────
def _sheet_public_ip(wb, credential, sub_id, region):
    ws = wb.create_sheet("PublicIP")
    _header(ws, ["Name","ResourceGroup","Location","IPAddress","AllocationMethod",
                 "SKU","Zones","AssociatedTo","SubscriptionId"])
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    for ip in client.public_ip_addresses.list_all():
        if not _match(ip, region): continue
        rg    = _rg(ip.id)
        sku   = str(ip.sku.name) if getattr(ip,"sku",None) else ""
        zones = ",".join(ip.zones) if getattr(ip,"zones",None) else ""
        assoc = ""
        if getattr(ip,"ip_configuration",None) and ip.ip_configuration.id:
            assoc = ip.ip_configuration.id.split("/")[-1]
        ws.append([ip.name, rg, ip.location, getattr(ip,"ip_address","") or "",
                   str(getattr(ip,"public_ip_allocation_method","") or ""),
                   sku, zones, assoc, sub_id])
    _autofit(ws)

# ─── Load Balancer (Network Flow: Overview → Rule → BackendPool → HealthProbe) ──
def _sheet_load_balancer(wb, credential, sub_id, region):
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    lbs = [lb for lb in client.load_balancers.list_all() if _match(lb, region)]

    # 시트1: LB Overview + Frontend IP
    ws1 = wb.create_sheet("LB_Overview")
    _header(ws1, ["LBName","ResourceGroup","Location","SKU","Tier","ProvisioningState",
                  "FE_Name","FE_PrivateIP","FE_PublicIP","FE_Subnet","FE_Zones",
                  "→ Rules","→ BackendPools","→ HealthProbes"], color=C_BLUE)
    for lb in lbs:
        rg   = _rg(lb.id)
        sku  = str(lb.sku.name) if getattr(lb,"sku",None) else ""
        tier = str(lb.sku.tier) if getattr(lb,"sku",None) and lb.sku.tier else ""
        fes  = lb.frontend_ip_configurations or []
        for idx, fe in enumerate(fes or [None]):
            rn = ws1.max_row + 1
            if fe:
                pub    = fe.public_ip_address.id.split("/")[-1] if getattr(fe,"public_ip_address",None) and fe.public_ip_address.id else ""
                subnet = fe.subnet.id.split("/")[-1] if getattr(fe,"subnet",None) and fe.subnet.id else ""
                zones  = ",".join(fe.zones) if getattr(fe,"zones",None) else ""
                fv = [fe.name or "", getattr(fe,"private_ip_address","") or "", pub, subnet, zones]
            else:
                fv = ["","","","",""]
            ws1.append([lb.name, rg, lb.location, sku, tier,
                        getattr(lb,"provisioning_state","") or ""] + fv + ["","",""])
            if idx == 0:
                _link(ws1, rn, 12, "LB_Rules",        "→ Rules")
                _link(ws1, rn, 13, "LB_BackendPools",  "→ Pools")
                _link(ws1, rn, 14, "LB_HealthProbes",  "→ Probes")
    _autofit(ws1)

    # 시트2: LB Rules
    ws2 = wb.create_sheet("LB_Rules")
    _header(ws2, ["LBName","ResourceGroup","RuleName","Protocol",
                  "FrontendPort","BackendPort","FrontendIPConfig",
                  "→ BackendPool","→ HealthProbe",
                  "LoadDistribution","IdleTimeoutMin","FloatingIP","TCPReset",
                  "← Overview"], color=C_ORANGE)
    for lb in lbs:
        rg = _rg(lb.id)
        for rule in (lb.load_balancing_rules or []):
            rn   = ws2.max_row + 1
            fe_n = rule.frontend_ip_configuration.id.split("/")[-1] if getattr(rule,"frontend_ip_configuration",None) else ""
            be_n = rule.backend_address_pool.id.split("/")[-1] if getattr(rule,"backend_address_pool",None) else ""
            pr_n = rule.probe.id.split("/")[-1] if getattr(rule,"probe",None) else ""
            ws2.append([lb.name, rg, rule.name or "",
                        str(getattr(rule,"protocol","") or ""),
                        getattr(rule,"frontend_port","") or "",
                        getattr(rule,"backend_port","") or "",
                        fe_n, be_n, pr_n,
                        str(getattr(rule,"load_distribution","") or ""),
                        getattr(rule,"idle_timeout_in_minutes","") or "",
                        str(getattr(rule,"enable_floating_ip","") or ""),
                        str(getattr(rule,"enable_tcp_reset","") or ""), ""])
            _link(ws2, rn, 8,  "LB_BackendPools", be_n or "→ Pool")
            _link(ws2, rn, 9,  "LB_HealthProbes", pr_n or "→ Probe")
            _link(ws2, rn, 14, "LB_Overview",     "← Overview")
    _autofit(ws2)

    # 시트3: Backend Pools
    ws3 = wb.create_sheet("LB_BackendPools")
    _header(ws3, ["LBName","ResourceGroup","PoolName","IPCount","BackendIPs","← Rules"], color=C_TEAL)
    for lb in lbs:
        rg = _rg(lb.id)
        for bp in (lb.backend_address_pools or []):
            rn  = ws3.max_row + 1
            ips = [getattr(a,"ip_address","") for a in (getattr(bp,"load_balancer_backend_addresses",None) or []) if getattr(a,"ip_address",None)]
            ws3.append([lb.name, rg, bp.name or "", len(ips), ",".join(ips), ""])
            _link(ws3, rn, 6, "LB_Rules", "← Rules")
    _autofit(ws3)

    # 시트4: Health Probes
    ws4 = wb.create_sheet("LB_HealthProbes")
    _header(ws4, ["LBName","ResourceGroup","ProbeName","Protocol","Port",
                  "IntervalSec","UnhealthyThreshold","RequestPath","← Rules"], color=C_GRAY)
    for lb in lbs:
        rg = _rg(lb.id)
        for p in (lb.probes or []):
            rn = ws4.max_row + 1
            ws4.append([lb.name, rg, p.name or "",
                        str(getattr(p,"protocol","") or ""),
                        getattr(p,"port","") or "",
                        getattr(p,"interval_in_seconds","") or "",
                        getattr(p,"number_of_probes","") or "",
                        getattr(p,"request_path","") or "", ""])
            _link(ws4, rn, 9, "LB_Rules", "← Rules")
    _autofit(ws4)

# ─── App Gateway (Network Flow: Overview → Listener → Rule → BackendPool/Settings → HealthProbe) ──
def _sheet_app_gateway(wb, credential, sub_id, region):
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    agws = [a for a in client.application_gateways.list_all() if _match(a, region)]

    # 시트1: AGW Overview + Frontend IP
    ws1 = wb.create_sheet("AGW_Overview")
    _header(ws1, ["AGWName","ResourceGroup","Location","SKU","Tier",
                  "CapacityMin","CapacityMax","OperationalState","ProvisioningState",
                  "FE_Name","FE_PrivateIP","FE_PublicIP",
                  "→ Listeners","→ Rules","→ Pools","→ Settings","→ Probes"], color=C_BLUE)
    for agw in agws:
        rg   = _rg(agw.id)
        sku  = str(agw.sku.name) if getattr(agw,"sku",None) else ""
        tier = str(agw.sku.tier) if getattr(agw,"sku",None) and agw.sku.tier else ""
        cmin = cmax = ""
        if getattr(agw,"autoscale_configuration",None):
            cmin = getattr(agw.autoscale_configuration,"min_capacity","") or ""
            cmax = getattr(agw.autoscale_configuration,"max_capacity","") or ""
        elif getattr(agw,"sku",None) and agw.sku.capacity:
            cmin = cmax = agw.sku.capacity
        for idx, fe in enumerate(agw.frontend_ip_configurations or [None]):
            rn = ws1.max_row + 1
            if fe:
                pub = fe.public_ip_address.id.split("/")[-1] if getattr(fe,"public_ip_address",None) and fe.public_ip_address.id else ""
                fv  = [fe.name or "", getattr(fe,"private_ip_address","") or "", pub]
            else:
                fv = ["","",""]
            ws1.append([agw.name, rg, agw.location, sku, tier, cmin, cmax,
                        str(getattr(agw,"operational_state","") or ""),
                        str(getattr(agw,"provisioning_state","") or "")] + fv + ["","","","",""])
            if idx == 0:
                _link(ws1, rn, 13, "AGW_Listeners",       "→ Listeners")
                _link(ws1, rn, 14, "AGW_Rules",           "→ Rules")
                _link(ws1, rn, 15, "AGW_BackendPools",    "→ Pools")
                _link(ws1, rn, 16, "AGW_BackendSettings", "→ Settings")
                _link(ws1, rn, 17, "AGW_HealthProbes",    "→ Probes")
    _autofit(ws1)

    # 시트2: Listeners
    ws2 = wb.create_sheet("AGW_Listeners")
    _header(ws2, ["AGWName","ResourceGroup","ListenerName","Protocol",
                  "FrontendIP","FrontendPort","HostName","HostNames",
                  "RequireSNI","SslCert","→ Rules","← Overview"], color=C_INDIGO)
    for agw in agws:
        rg = _rg(agw.id)
        for lst in (agw.http_listeners or []):
            rn      = ws2.max_row + 1
            fe_name = lst.frontend_ip_configuration.id.split("/")[-1] if getattr(lst,"frontend_ip_configuration",None) else ""
            port_nm = lst.frontend_port.id.split("/")[-1] if getattr(lst,"frontend_port",None) else ""
            ssl_nm  = lst.ssl_certificate.id.split("/")[-1] if getattr(lst,"ssl_certificate",None) else ""
            hnames  = ",".join(getattr(lst,"host_names",None) or [])
            ws2.append([agw.name, rg, lst.name or "",
                        str(getattr(lst,"protocol","") or ""),
                        fe_name, port_nm,
                        getattr(lst,"host_name","") or "", hnames,
                        str(getattr(lst,"require_server_name_indication","") or ""),
                        ssl_nm, "", ""])
            _link(ws2, rn, 11, "AGW_Rules",    "→ Rules")
            _link(ws2, rn, 12, "AGW_Overview", "← Overview")
    _autofit(ws2)

    # 시트3: Rules
    ws3 = wb.create_sheet("AGW_Rules")
    _header(ws3, ["AGWName","ResourceGroup","RuleName","RuleType",
                  "→ Listener","→ BackendPool","→ BackendSettings",
                  "RedirectConfig","RewriteRuleSet","← Overview"], color=C_ORANGE)
    for agw in agws:
        rg = _rg(agw.id)
        for rule in (agw.request_routing_rules or []):
            rn       = ws3.max_row + 1
            listener = rule.http_listener.id.split("/")[-1] if getattr(rule,"http_listener",None) else ""
            be_pool  = rule.backend_address_pool.id.split("/")[-1] if getattr(rule,"backend_address_pool",None) else ""
            be_set   = rule.backend_http_settings.id.split("/")[-1] if getattr(rule,"backend_http_settings",None) else ""
            redirect = rule.redirect_configuration.id.split("/")[-1] if getattr(rule,"redirect_configuration",None) else ""
            rewrite  = rule.rewrite_rule_set.id.split("/")[-1] if getattr(rule,"rewrite_rule_set",None) else ""
            ws3.append([agw.name, rg, rule.name or "",
                        str(getattr(rule,"rule_type","") or ""),
                        listener, be_pool, be_set, redirect, rewrite, ""])
            _link(ws3, rn, 5,  "AGW_Listeners",       listener or "→ Listener")
            _link(ws3, rn, 6,  "AGW_BackendPools",    be_pool or "→ Pool")
            _link(ws3, rn, 7,  "AGW_BackendSettings", be_set or "→ Settings")
            _link(ws3, rn, 10, "AGW_Overview",        "← Overview")
    _autofit(ws3)

    # 시트4: Backend Settings
    ws4 = wb.create_sheet("AGW_BackendSettings")
    _header(ws4, ["AGWName","ResourceGroup","SettingName","Protocol","Port",
                  "CookieAffinity","RequestTimeout","PickHostFromBackend","HostName",
                  "→ HealthProbe","← Rules"], color=C_TEAL)
    for agw in agws:
        rg = _rg(agw.id)
        for hs in (agw.backend_http_settings_collection or []):
            rn    = ws4.max_row + 1
            probe = hs.probe.id.split("/")[-1] if getattr(hs,"probe",None) else ""
            ws4.append([agw.name, rg, hs.name or "",
                        str(getattr(hs,"protocol","") or ""),
                        getattr(hs,"port","") or "",
                        str(getattr(hs,"cookie_based_affinity","") or ""),
                        getattr(hs,"request_timeout","") or "",
                        str(getattr(hs,"pick_host_name_from_backend_address","") or ""),
                        getattr(hs,"host_name","") or "", probe, ""])
            _link(ws4, rn, 10, "AGW_HealthProbes", probe or "→ Probe")
            _link(ws4, rn, 11, "AGW_Rules",        "← Rules")
    _autofit(ws4)

    # 시트5: Backend Pools
    ws5 = wb.create_sheet("AGW_BackendPools")
    _header(ws5, ["AGWName","ResourceGroup","PoolName","AddrCount","Addresses","← Rules"], color=C_TEAL)
    for agw in agws:
        rg = _rg(agw.id)
        for pool in (agw.backend_address_pools or []):
            rn    = ws5.max_row + 1
            addrs = [getattr(a,"ip_address",None) or getattr(a,"fqdn",None) or "" for a in (getattr(pool,"backend_addresses",None) or [])]
            addrs = [a for a in addrs if a]
            ws5.append([agw.name, rg, pool.name or "", len(addrs), ",".join(addrs), ""])
            _link(ws5, rn, 6, "AGW_Rules", "← Rules")
    _autofit(ws5)

    # 시트6: Health Probes
    ws6 = wb.create_sheet("AGW_HealthProbes")
    _header(ws6, ["AGWName","ResourceGroup","ProbeName","Protocol","Host","Path",
                  "IntervalSec","TimeoutSec","UnhealthyThreshold","PickHostFromBackend",
                  "MinServers","← BackendSettings"], color=C_GRAY)
    for agw in agws:
        rg = _rg(agw.id)
        for p in (agw.probes or []):
            rn = ws6.max_row + 1
            ws6.append([agw.name, rg, p.name or "",
                        str(getattr(p,"protocol","") or ""),
                        getattr(p,"host","") or "",
                        getattr(p,"path","") or "",
                        getattr(p,"interval","") or "",
                        getattr(p,"timeout","") or "",
                        getattr(p,"unhealthy_threshold","") or "",
                        str(getattr(p,"pick_host_name_from_backend_http_settings","") or ""),
                        getattr(p,"min_servers","") or "", ""])
            _link(ws6, rn, 12, "AGW_BackendSettings", "← Settings")
    _autofit(ws6)

# ─── Virtual WAN / Hub / VPN GW / Express Route ───────────────────────
def _sheet_virtual_wan(wb, credential, sub_id, region):
    ws = wb.create_sheet("VirtualWAN")
    _header(ws, ["Name","ResourceGroup","Location","Type","ProvisioningState","SubscriptionId"])
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.virtual_wans.list():
        if not _match(item, region): continue
        ws.append([item.name, _rg(item.id), item.location,
                   getattr(item,"virtual_wan_type","") or "",
                   getattr(item,"provisioning_state","") or "", sub_id])
    _autofit(ws)

def _sheet_virtual_hub(wb, credential, sub_id, region):
    ws = wb.create_sheet("VirtualHub")
    _header(ws, ["Name","ResourceGroup","Location","AddressPrefix","VirtualWAN","SKU","ProvisioningState","SubscriptionId"])
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.virtual_hubs.list():
        if not _match(item, region): continue
        vwan = item.virtual_wan.id.split("/")[-1] if getattr(item,"virtual_wan",None) else ""
        ws.append([item.name, _rg(item.id), item.location,
                   getattr(item,"address_prefix","") or "", vwan,
                   getattr(item,"sku","") or "",
                   getattr(item,"provisioning_state","") or "", sub_id])
    _autofit(ws)

def _sheet_vpn_gateway(wb, credential, sub_id, region):
    ws = wb.create_sheet("VPNGateway")
    _header(ws, ["Name","ResourceGroup","Location","VirtualHub","BgpPeeringAddr","ProvisioningState","SubscriptionId"])
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.vpn_gateways.list():
        if not _match(item, region): continue
        hub = item.virtual_hub.id.split("/")[-1] if getattr(item,"virtual_hub",None) else ""
        bgp = str(getattr(item.bgp_settings,"bgp_peering_address","") or "") if getattr(item,"bgp_settings",None) else ""
        ws.append([item.name, _rg(item.id), item.location, hub, bgp,
                   getattr(item,"provisioning_state","") or "", sub_id])
    _autofit(ws)

def _sheet_express_route(wb, credential, sub_id, region):
    ws = wb.create_sheet("ExpressRoute")
    _header(ws, ["Name","ResourceGroup","Location","ServiceProvider","PeeringLocation",
                 "BandwidthMbps","SKU","Tier","ProvisioningState","SubscriptionId"])
    client = NetworkManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.express_route_circuits.list_all():
        if not _match(item, region): continue
        sp  = item.service_provider_properties
        sku = item.sku if getattr(item,"sku",None) else None
        ws.append([item.name, _rg(item.id), item.location,
                   sp.service_provider_name if sp else "",
                   sp.peering_location if sp else "",
                   sp.bandwidth_in_mbps if sp else "",
                   str(sku.name) if sku else "",
                   str(sku.tier) if sku else "",
                   getattr(item,"circuit_provisioning_state","") or "", sub_id])
    _autofit(ws)

# ─── VM ───────────────────────────────────────────────────────────────
def _sheet_vm(wb, credential, sub_id, region):
    ws = wb.create_sheet("VM")
    _header(ws, ["Name","ResourceGroup","Location","VMSize","OSType","OSDiskSizeGB","SubscriptionId"])
    client = ComputeManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.virtual_machines.list_all():
        if not _match(item, region): continue
        rg     = _rg(item.id)
        hw     = item.hardware_profile
        sp     = getattr(item,"storage_profile",None)
        osd    = getattr(sp,"os_disk",None) if sp else None
        ws.append([item.name, rg, item.location,
                   hw.vm_size if hw else "",
                   str(getattr(osd,"os_type","") or "") if osd else "",
                   getattr(osd,"disk_size_gb","") or "" if osd else "", sub_id])
    _autofit(ws)

# ─── Container ────────────────────────────────────────────────────────
def _sheet_cae(wb, credential, sub_id, region):
    ws = wb.create_sheet("CAE")
    _header(ws, ["Name","ResourceGroup","Location","DefaultDomain","StaticIP","ZoneRedundant","SubscriptionId"])
    client = ContainerAppsAPIClient(credential=credential, subscription_id=sub_id)
    for item in client.managed_environments.list_by_subscription():
        if not _match(item, region): continue
        ws.append([item.name, _rg(item.id), item.location,
                   getattr(item,"default_domain","") or "",
                   getattr(item,"static_ip","") or "",
                   str(getattr(item,"zone_redundant","") or ""), sub_id])
    _autofit(ws)

def _sheet_aca(wb, credential, sub_id, region):
    ws = wb.create_sheet("ACA")
    _header(ws, ["Name","ResourceGroup","Location","ManagedEnv","IngressEnabled",
                 "IngressFQDN","External","ProvisioningState","SubscriptionId"])
    client = ContainerAppsAPIClient(credential=credential, subscription_id=sub_id)
    for item in client.container_apps.list_by_subscription():
        if not _match(item, region): continue
        env     = item.managed_environment_id.split("/")[-1] if getattr(item,"managed_environment_id",None) else ""
        ingress = getattr(item,"configuration",None) and getattr(item.configuration,"ingress",None)
        ws.append([item.name, _rg(item.id), item.location, env,
                   str(bool(ingress)),
                   getattr(ingress,"fqdn","") or "" if ingress else "",
                   str(getattr(ingress,"external","") or "") if ingress else "",
                   getattr(item,"provisioning_state","") or "", sub_id])
    _autofit(ws)

def _sheet_acr(wb, credential, sub_id, region):
    ws = wb.create_sheet("ACR")
    _header(ws, ["Name","ResourceGroup","Location","SKU","LoginServer","AdminEnabled","ProvisioningState","SubscriptionId"])
    client = ContainerRegistryManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.registries.list():
        if not _match(item, region): continue
        ws.append([item.name, _rg(item.id), item.location,
                   str(item.sku.name) if getattr(item,"sku",None) else "",
                   getattr(item,"login_server","") or "",
                   str(getattr(item,"admin_user_enabled","") or ""),
                   getattr(item,"provisioning_state","") or "", sub_id])
    _autofit(ws)

# ─── DB ───────────────────────────────────────────────────────────────
def _sheet_mongodb(wb, credential, sub_id, region):
    ws = wb.create_sheet("MongoDB")
    _header(ws, ["Name","ResourceGroup","Location","Kind","ConsistencyLevel","EnableFreeTier","ProvisioningState","SubscriptionId"])
    client = CosmosDBManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.database_accounts.list():
        if not _match(item, region): continue
        caps = [c.name for c in (item.capabilities or [])]
        if getattr(item,"kind",None) != "MongoDB" and "EnableMongo" not in caps: continue
        con = str(item.consistency_policy.default_consistency_level or "") if getattr(item,"consistency_policy",None) else ""
        ws.append([item.name, _rg(item.id), item.location,
                   getattr(item,"kind","") or "", con,
                   str(getattr(item,"enable_free_tier","") or ""),
                   getattr(item,"provisioning_state","") or "", sub_id])
    _autofit(ws)

def _sheet_postgresql(wb, credential, sub_id, region):
    ws = wb.create_sheet("PostgreSQL")
    _header(ws, ["Name","ResourceGroup","Location","Version","SKU","StorageGB","HAMode","State","SubscriptionId"])
    client = PostgreSQLManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.servers.list():
        if not _match(item, region): continue
        sku  = str(item.sku.name) if getattr(item,"sku",None) else ""
        stor = getattr(item,"storage",None)
        ha   = getattr(item,"high_availability",None)
        ws.append([item.name, _rg(item.id), item.location,
                   str(getattr(item,"version","") or ""), sku,
                   getattr(stor,"storage_size_gb","") or "" if stor else "",
                   str(getattr(ha,"mode","") or "") if ha else "",
                   str(getattr(item,"state","") or ""), sub_id])
    _autofit(ws)

# ─── Redis (신규) ─────────────────────────────────────────────────────
def _sheet_redis(wb, credential, sub_id, region):
    ws = wb.create_sheet("Redis")
    _header(ws, ["Name","ResourceGroup","Location","SKU","Family","Capacity",
                 "EnableNonSslPort","MinTlsVersion","RedisVersion",
                 "HostName","Port","SslPort","ProvisioningState","SubscriptionId"], color=C_ORANGE)
    client = RedisManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.redis.list_by_subscription():
        if not _match(item, region): continue
        sku = getattr(item,"sku",None)
        ws.append([item.name, _rg(item.id), item.location,
                   str(getattr(sku,"name","") or "") if sku else "",
                   str(getattr(sku,"family","") or "") if sku else "",
                   getattr(sku,"capacity","") or "" if sku else "",
                   str(getattr(item,"enable_non_ssl_port","") or ""),
                   str(getattr(item,"minimum_tls_version","") or ""),
                   getattr(item,"redis_version","") or "",
                   getattr(item,"host_name","") or "",
                   getattr(item,"port","") or "",
                   getattr(item,"ssl_port","") or "",
                   getattr(item,"provisioning_state","") or "", sub_id])
    _autofit(ws)

# ─── AKS (신규, 2시트) ────────────────────────────────────────────────
def _sheet_aks(wb, credential, sub_id, region):
    client   = ContainerServiceClient(credential=credential, subscription_id=sub_id)
    clusters = [c for c in client.managed_clusters.list() if _match(c, region)]

    ws1 = wb.create_sheet("AKS_Clusters")
    _header(ws1, ["Name","ResourceGroup","Location","K8sVersion","DNSPrefix","FQDN",
                  "NetworkPlugin","NetworkPolicy","PodCIDR","ServiceCIDR",
                  "ProvisioningState","PowerState","SubscriptionId","→ NodePools"], color=C_TEAL)
    for cl in clusters:
        rn    = ws1.max_row + 1
        net   = getattr(cl,"network_profile",None)
        power = getattr(cl,"power_state",None)
        ws1.append([cl.name, _rg(cl.id), cl.location,
                    getattr(cl,"kubernetes_version","") or "",
                    getattr(cl,"dns_prefix","") or "",
                    getattr(cl,"fqdn","") or "",
                    str(getattr(net,"network_plugin","") or "") if net else "",
                    str(getattr(net,"network_policy","") or "") if net else "",
                    getattr(net,"pod_cidr","") or "" if net else "",
                    getattr(net,"service_cidr","") or "" if net else "",
                    getattr(cl,"provisioning_state","") or "",
                    str(getattr(power,"code","") or "") if power else "",
                    sub_id, ""])
        _link(ws1, rn, 14, "AKS_NodePools", "→ NodePools")
    _autofit(ws1)

    ws2 = wb.create_sheet("AKS_NodePools")
    _header(ws2, ["ClusterName","ResourceGroup","PoolName","Mode","OSType","VMSize",
                  "NodeCount","MinCount","MaxCount","AutoScaling",
                  "OSDiskGB","MaxPods","K8sVersion","ProvisioningState","← Clusters"], color=C_TEAL)
    for cl in clusters:
        rg = _rg(cl.id)
        for pool in (getattr(cl,"agent_pool_profiles",None) or []):
            rn = ws2.max_row + 1
            ws2.append([cl.name, rg, pool.name or "",
                        str(getattr(pool,"mode","") or ""),
                        str(getattr(pool,"os_type","") or ""),
                        getattr(pool,"vm_size","") or "",
                        getattr(pool,"count","") or "",
                        getattr(pool,"min_count","") or "",
                        getattr(pool,"max_count","") or "",
                        str(getattr(pool,"enable_auto_scaling","") or ""),
                        getattr(pool,"os_disk_size_gb","") or "",
                        getattr(pool,"max_pods","") or "",
                        getattr(pool,"orchestrator_version","") or "",
                        getattr(pool,"provisioning_state","") or "", ""])
            _link(ws2, rn, 15, "AKS_Clusters", "← Clusters")
    _autofit(ws2)

# ─── AI Search / AOAI ─────────────────────────────────────────────────
def _sheet_ai_search(wb, credential, sub_id, region):
    ws = wb.create_sheet("AISearch")
    _header(ws, ["Name","ResourceGroup","Location","SKU","ReplicaCount","PartitionCount","Status","SubscriptionId"])
    client = SearchManagementClient(credential=credential, subscription_id=sub_id)
    for item in client.services.list_by_subscription():
        if not _match(item, region): continue
        ws.append([item.name, _rg(item.id), item.location,
                   str(item.sku.name) if getattr(item,"sku",None) else "",
                   getattr(item,"replica_count","") or "",
                   getattr(item,"partition_count","") or "",
                   str(getattr(item,"status","") or ""), sub_id])
    _autofit(ws)

def _sheet_aoai(wb, credential, sub_id, region):
    client = CognitiveServicesManagementClient(credential=credential, subscription_id=sub_id)
    accounts = [a for a in client.accounts.list()
                if _match(a, region) and getattr(a,"kind",None) == "OpenAI"]

    ws1 = wb.create_sheet("AOAI_Accounts")
    _header(ws1, ["Name","ResourceGroup","Location","SKU","CustomSubDomain",
                  "PublicNetworkAccess","ProvisioningState","SubscriptionId","→ Models"])
    for acc in accounts:
        rn    = ws1.max_row + 1
        sku   = str(acc.sku.name) if getattr(acc,"sku",None) else ""
        props = getattr(acc,"properties",None)
        ws1.append([acc.name, _rg(acc.id), acc.location, sku,
                    getattr(props,"custom_sub_domain_name","") or "" if props else "",
                    str(getattr(props,"public_network_access","") or "") if props else "",
                    getattr(props,"provisioning_state","") or "" if props else "",
                    sub_id, ""])
        _link(ws1, rn, 9, "AOAI_Models", "→ Models")
    _autofit(ws1)

    ws2 = wb.create_sheet("AOAI_Models")
    _header(ws2, ["AccountName","ResourceGroup","DeploymentName","ModelName","ModelVersion",
                  "ModelFormat","SKU","Capacity","ProvisioningState","← Accounts"])
    for acc in accounts:
        rg = _rg(acc.id)
        try:
            for dep in client.deployments.list(rg, acc.name):
                rn    = ws2.max_row + 1
                model = getattr(dep,"model",None)
                dsku  = getattr(dep,"sku",None)
                ws2.append([acc.name, rg, dep.name or "",
                            getattr(model,"name","") or "" if model else "",
                            getattr(model,"version","") or "" if model else "",
                            getattr(model,"format","") or "" if model else "",
                            str(getattr(dsku,"name","") or "") if dsku else "",
                            getattr(dsku,"capacity","") or "" if dsku else "",
                            getattr(dep,"provisioning_state","") or "", ""])
                _link(ws2, rn, 10, "AOAI_Accounts", "← Accounts")
        except Exception:
            pass
    _autofit(ws2)

# ─── AI Foundry (Hub / Project / 배포 모델) ──────────────────────────
# ─── Foundry (AIServices) ────────────────────────────────────────────
def _sheet_foundry(wb, credential, sub_id, region):
    from azure.mgmt.resource import ResourceManagementClient
    
    resource_client = ResourceManagementClient(credential=credential, subscription_id=sub_id)
    cognitive_client = CognitiveServicesManagementClient(credential=credential, subscription_id=sub_id)
    
    ws = wb.create_sheet("Foundry")
    _header(ws, [
        "Name", "ResourceGroup", "Location", "SKU", "Kind",
        "PublicNetworkAccess", "ProvisioningState", "SubscriptionId"
    ], color=C_INDIGO)
    
    try:
        resources = resource_client.resources.list(
            filter="resourceType eq 'Microsoft.CognitiveServices/accounts'"
        )
        
        for resource in resources:
            if not _match(resource, region):
                continue
            
            rg = resource.id.split('/')[4]
            
            try:
                account = cognitive_client.accounts.get(rg, resource.name)
                kind = str(getattr(account, "kind", "") or "")
                
                # AIServices만 필터링
                if kind != "AIServices":
                    continue
                
                props = getattr(account, "properties", None)
                pna = str(getattr(props, "public_network_access", "") or "") if props else ""
                ps = str(getattr(props, "provisioning_state", "") or "") if props else ""
                
                ws.append([
                    account.name or "",
                    rg,
                    getattr(account, "location", "") or "",
                    str(account.sku.name) if getattr(account, "sku", None) else "",
                    kind,
                    pna,
                    ps,
                    sub_id
                ])
            except Exception:
                continue
    except Exception as e:
        ws.append(["조회 오류", str(e)])
    
    _autofit(ws)

# ─── Speech Services ──────────────────────────────────────────────────
def _sheet_speech(wb, credential, sub_id, region):
    from azure.mgmt.resource import ResourceManagementClient
    
    resource_client = ResourceManagementClient(credential=credential, subscription_id=sub_id)
    cognitive_client = CognitiveServicesManagementClient(credential=credential, subscription_id=sub_id)
    
    ws = wb.create_sheet("SpeechServices")
    _header(ws, [
        "Name", "ResourceGroup", "Location", "SKU", "Kind",
        "PublicNetworkAccess", "ProvisioningState", "SubscriptionId"
    ], color=C_INDIGO)
    
    try:
        resources = resource_client.resources.list(
            filter="resourceType eq 'Microsoft.CognitiveServices/accounts'"
        )
        
        for resource in resources:
            if not _match(resource, region):
                continue
            
            rg = resource.id.split('/')[4]
            
            try:
                account = cognitive_client.accounts.get(rg, resource.name)
                kind = str(getattr(account, "kind", "") or "")
                
                # SpeechServices만 필터링
                if kind != "SpeechServices":
                    continue
                
                props = getattr(account, "properties", None)
                pna = str(getattr(props, "public_network_access", "") or "") if props else ""
                ps = str(getattr(props, "provisioning_state", "") or "") if props else ""
                
                ws.append([
                    account.name or "",
                    rg,
                    getattr(account, "location", "") or "",
                    str(account.sku.name) if getattr(account, "sku", None) else "",
                    kind,
                    pna,
                    ps,
                    sub_id
                ])
            except Exception:
                continue
    except Exception as e:
        ws.append(["조회 오류", str(e)])
    
    _autofit(ws)

# ─── Form Recognizer ──────────────────────────────────────────────────
def _sheet_form_recognizer(wb, credential, sub_id, region):
    from azure.mgmt.resource import ResourceManagementClient
    
    resource_client = ResourceManagementClient(credential=credential, subscription_id=sub_id)
    cognitive_client = CognitiveServicesManagementClient(credential=credential, subscription_id=sub_id)
    
    ws = wb.create_sheet("FormRecognizer")
    _header(ws, [
        "Name", "ResourceGroup", "Location", "SKU", "Kind",
        "PublicNetworkAccess", "ProvisioningState", "SubscriptionId"
    ], color=C_INDIGO)
    
    try:
        resources = resource_client.resources.list(
            filter="resourceType eq 'Microsoft.CognitiveServices/accounts'"
        )
        
        for resource in resources:
            if not _match(resource, region):
                continue
            
            rg = resource.id.split('/')[4]
            
            try:
                account = cognitive_client.accounts.get(rg, resource.name)
                kind = str(getattr(account, "kind", "") or "")
                
                # FormRecognizer만 필터링
                if kind != "FormRecognizer":
                    continue
                
                props = getattr(account, "properties", None)
                pna = str(getattr(props, "public_network_access", "") or "") if props else ""
                ps = str(getattr(props, "provisioning_state", "") or "") if props else ""
                
                ws.append([
                    account.name or "",
                    rg,
                    getattr(account, "location", "") or "",
                    str(account.sku.name) if getattr(account, "sku", None) else "",
                    kind,
                    pna,
                    ps,
                    sub_id
                ])
            except Exception:
                continue
    except Exception as e:
        ws.append(["조회 오류", str(e)])
    
    _autofit(ws)

# ─── AI Foundry (이전 버전 - 호환성 유지) ─────────────────────────────
def _sheet_ai_foundry(wb, credential, sub_id, region):
    try:
        from azure.mgmt.machinelearningservices import MachineLearningServicesMgmtClient
        client = MachineLearningServicesMgmtClient(credential=credential, subscription_id=sub_id)
    except ImportError:
        ws = wb.create_sheet("AIFoundry_ERROR")
        ws.append(["패키지 필요", "pip install azure-mgmt-machinelearningservices"])
        return

    # ── 시트1: Hubs ───────────────────────────────────────────────────
    ws_hub = wb.create_sheet("AIFoundry_Hubs")
    _header(ws_hub, [
        "HubName", "ResourceGroup", "Location", "SKU", "FriendlyName",
        "PublicNetworkAccess", "ProvisioningState", "SubscriptionId",
        "→ Projects",
    ], color=C_INDIGO)

    hubs = []
    try:
        for ws_obj in client.workspaces.list_by_subscription():
            if not _match(ws_obj, region): continue
            if str(getattr(ws_obj, "kind", "") or "").lower() != "hub": continue
            hubs.append(ws_obj)
            rn  = ws_hub.max_row + 1
            sku = str(ws_obj.sku.name) if getattr(ws_obj, "sku", None) else ""
            ws_hub.append([
                ws_obj.name or "", _rg(ws_obj.id), ws_obj.location or "",
                sku,
                str(getattr(ws_obj, "friendly_name", "") or ""),
                str(getattr(ws_obj, "public_network_access", "") or ""),
                str(getattr(ws_obj, "provisioning_state", "") or ""),
                sub_id, "",
            ])
            _link(ws_hub, rn, 9, "AIFoundry_Projects", "→ Projects")
    except Exception as e:
        ws_hub.append(["조회 오류", str(e)])
    _autofit(ws_hub)

    # ── 시트2: Projects ───────────────────────────────────────────────
    ws_proj = wb.create_sheet("AIFoundry_Projects")
    _header(ws_proj, [
        "ProjectName", "ResourceGroup", "Location", "HubName", "FriendlyName",
        "PublicNetworkAccess", "ProvisioningState", "SubscriptionId",
        "← Hubs", "→ Models",
    ], color=C_INDIGO)

    projects = []
    try:
        for ws_obj in client.workspaces.list_by_subscription():
            if not _match(ws_obj, region): continue
            if str(getattr(ws_obj, "kind", "") or "").lower() != "project": continue
            projects.append(ws_obj)
            rn       = ws_proj.max_row + 1
            hub_name = ""
            if getattr(ws_obj, "hub_resource_id", None):
                hub_name = ws_obj.hub_resource_id.split("/")[-1]
            ws_proj.append([
                ws_obj.name or "", _rg(ws_obj.id), ws_obj.location or "",
                hub_name,
                str(getattr(ws_obj, "friendly_name", "") or ""),
                str(getattr(ws_obj, "public_network_access", "") or ""),
                str(getattr(ws_obj, "provisioning_state", "") or ""),
                sub_id, "", "",
            ])
            _link(ws_proj, rn, 9,  "AIFoundry_Hubs",   "← Hubs")
            _link(ws_proj, rn, 10, "AIFoundry_Models",  "→ Models")
    except Exception as e:
        ws_proj.append(["조회 오류", str(e)])
    _autofit(ws_proj)

    # ── 시트3: 배포 모델 ──────────────────────────────────────────────
    ws_model = wb.create_sheet("AIFoundry_Models")
    _header(ws_model, [
        "ProjectName", "ResourceGroup", "EndpointName", "DeploymentName",
        "ModelName", "ModelVersion", "ModelFormat",
        "InstanceType", "ProvisioningState",
        "← Projects",
    ], color=C_INDIGO)

    for proj in projects:
        rg = _rg(proj.id)
        try:
            for ep in client.online_endpoints.list(rg, proj.name):
                ep_name = ep.name or ""
                try:
                    for dep in client.online_deployments.list(rg, proj.name, ep_name):
                        rn    = ws_model.max_row + 1
                        model = getattr(dep, "model", None)
                        # ARM ID 문자열 또는 객체 모두 처리
                        if isinstance(model, str):
                            parts    = model.split("/")
                            try:    m_name = parts[parts.index("models") + 1]
                            except: m_name = parts[-1]
                            try:    m_ver  = parts[parts.index("versions") + 1]
                            except: m_ver  = ""
                            m_fmt = ""
                        elif model:
                            m_name = str(getattr(model, "name", "")    or getattr(model, "model_name", "")    or "")
                            m_ver  = str(getattr(model, "version", "") or getattr(model, "model_version", "") or "")
                            m_fmt  = str(getattr(model, "format", "")  or getattr(model, "model_format", "")  or "")
                        else:
                            m_name = m_ver = m_fmt = ""

                        ws_model.append([
                            proj.name or "", rg, ep_name, dep.name or "",
                            m_name, m_ver, m_fmt,
                            str(getattr(dep, "instance_type", "") or ""),
                            str(getattr(dep, "provisioning_state", "") or ""),
                            "",
                        ])
                        _link(ws_model, rn, 10, "AIFoundry_Projects", "← Projects")
                except Exception:
                    continue
        except Exception:
            continue
    _autofit(ws_model)


# ─── SHEET_BUILDERS ───────────────────────────────────────────────────
SHEET_BUILDERS = {
    "vnet_subnet":    _sheet_vnet_subnet,
    "nsg":            _sheet_nsg,
    "route_table":    _sheet_route_table,
    "public_ip":      _sheet_public_ip,
    "load_balancer":  _sheet_load_balancer,
    "app_gateway":    _sheet_app_gateway,
    "virtual_wan":    _sheet_virtual_wan,
    "virtual_hub":    _sheet_virtual_hub,
    "vpn_gateway":    _sheet_vpn_gateway,
    "express_route":  _sheet_express_route,
    "vm":             _sheet_vm,
    "cae":            _sheet_cae,
    "aca":            _sheet_aca,
    "acr":            _sheet_acr,
    "mongodb":        _sheet_mongodb,
    "postgresql":     _sheet_postgresql,
    "redis":          _sheet_redis,
    "aks":            _sheet_aks,
    "ai_search":      _sheet_ai_search,
    "aoai":           _sheet_aoai,
    "foundry":        _sheet_foundry,
    "speech":         _sheet_speech,
    "form_recognizer": _sheet_form_recognizer,
    "ai_foundry":     _sheet_ai_foundry,  # 이전 버전 호환성
}

# ─── 엔드포인트 ───────────────────────────────────────────────────────
@router.post("/export/inventory")
def export_inventory(req: ExportRequest) -> Response:
    if not all([req.subscription_id, req.tenant_id, req.client_id, req.client_secret]):
        raise HTTPException(status_code=400, detail="인증 정보가 모두 필요합니다.")
    if not req.services:
        raise HTTPException(status_code=400, detail="서비스를 하나 이상 선택해주세요.")

    credential = get_credential(req.tenant_id, req.client_id, req.client_secret)
    wb = Workbook()
    wb.remove(wb.active)

    for svc_key in req.services:
        builder = SHEET_BUILDERS.get(svc_key)
        if builder:
            try:
                builder(wb, credential, req.subscription_id, req.region)
            except Exception as e:
                ws = wb.create_sheet(f"{svc_key}_ERROR")
                ws.append(["오류 발생", str(e)])

    if not wb.sheetnames:
        raise HTTPException(status_code=404, detail="조회된 데이터가 없습니다.")

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        content=stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="azure_inventory.xlsx"'},
    )